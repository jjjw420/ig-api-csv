"""
    Market sentiment to excell.

    Hannes Wagener - 2021
"""
import requests
import json
import sys
import datetime
import os.path
from openpyxl import load_workbook
from openpyxl import Workbook


# modify the following with your IG Credentials
IG_IDENTIFIER = "your IG user name"
IG_PASSWORD = "your IG password"
IG_API_KEY = "your IG API Key"

# the following list determines the order in the spreadsheet.  
# NOTE!! If this order changes - you must start a new sheet, however it is safe to append new markets at the end.
IG_SENTIMENT_MARKETS = "EURGBP,EURUSD,EURAUD,EURCAD,EURCHF,EURNZD,EURJPY,GBPUSD,GBPAUD,GBPNZD,GBPCAD,GBPJPY,GBPCHF,AUDUSD,AUDNZD,AUDCAD,AUDCHF,AUDJPY,NZDCAD,NZDCHF,NZDJPY,NZDUSD,CADCHF,CADJPY,CHFJPY,USDJPY,USDCHF,USDCAD,USDSGD"

IG_API_BASE_URL = "https://api.ig.com/gateway/deal/"

date_format = "%Y-%m-%d %H:%M:%S"
output_xlsx_file = "client_sentiment.xlsx"
client_sentiment_sheet = "Client Sentiment"

# create a map for market to column number - in case the api fails to return an entry?
market_col = {}
c = 2
for m in IG_SENTIMENT_MARKETS.split(","):
    market_col[m.strip()] = c
    c = c + 1

headers = {
    "Content-Type": "application/json", 
    "Accept": "application/json", 
    "X-IG-API-KEY": IG_API_KEY, 
    "Version": "2"
}

auth_body = {
    "identifier": IG_IDENTIFIER,
    "password": IG_PASSWORD,
    "encryptedPassword": None
}

s = requests.session()

#IG auth
print("Authenticating...")
s.headers.update(headers)
r = s.post(url=IG_API_BASE_URL + "session", json=auth_body)

if r.status_code != 200:
    print("Authentication failed!")
    print("Status Code: %i" % r.status_code)
    print("Content: %s" % r.content)
    sys.exit(8)

print("Authentication Successfull!")

s.headers.update({
    "CST": r.headers["CST"], 
    "X-SECURITY-TOKEN": r.headers["X-SECURITY-TOKEN"]
})

print("Getting Market Sentiment for %s" % (IG_SENTIMENT_MARKETS))

#IG Sentiment API
r = s.get(url=IG_API_BASE_URL + "clientsentiment", params={"marketIds": IG_SENTIMENT_MARKETS})

if r.status_code != 200:
    print("Error!")
    print("Status Code: %i" % r.status_code)
    print("Content: %s" % r.content)
    sys.exit(8)

print("API Request Ok!")

resp_obj = json.loads(r.content)

wb = None
ws = None
# check if exists.  load if it does create and write headers if not.
if os.path.exists(output_xlsx_file):
    wb = load_workbook(output_xlsx_file)
    ws = wb[client_sentiment_sheet]
else:
    wb = Workbook()
    ws = wb.create_sheet(client_sentiment_sheet, 0) 

#overwrite the header in case it changed.
c = 2
ws['A1'] = "Date" 
for m in IG_SENTIMENT_MARKETS.split(","):
    if m.strip() != "":
        ws.cell(row=1, column=c, value=m.strip())
        c = c + 1

#get current rundate as string
dte = datetime.datetime.now()
dte_str = dte.strftime(date_format)

if len(resp_obj) > 0:
    #insert a row at the top
    ws.insert_rows(2)
    #populate columns in new row
    for client_sentiment in resp_obj["clientSentiments"]:
        client_sentiment["date"] = dte_str
        ws.cell(row=2, column=1, value=dte_str)
        #get column from market map
        c = market_col[client_sentiment["marketId"]]
        ws.cell(row=2, column=c, value=client_sentiment["longPositionPercentage"])
    
# close session
s.close()
# save xlsx
wb.save(output_xlsx_file)

print("Done! See: %s" % output_xlsx_file)